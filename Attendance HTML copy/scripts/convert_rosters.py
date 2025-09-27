#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


STATUS_MAP = {
    'Present': 'present',
    'Online': 'online',
    'Excused': 'excused',
    'Tardy': 'tardy',
    'Absent': 'absent',
    'Early Leave': 'early_leave',
    'Very Early Leave': 'very_early_leave',
    'Non-service': 'non_service',
}


DAY_ABBR = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def slugify_id(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip('_')
    return f"p_{s}"


def event_id(label: str) -> str:
    s = label.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s


def dow_from_date(dt: datetime) -> int:
    # Monday=1 .. Sunday=7
    return (dt.isoweekday())


def parse_attendance_sheet(xlsx_path: Path) -> Dict:
    wb = load_workbook(xlsx_path, data_only=True)
    if 'Attendance' not in wb.sheetnames:
        raise ValueError('Missing Attendance sheet')
    ws = wb['Attendance']

    # Identify name columns from row 1 (skip instructions/noise)
    name_cols: List[Tuple[int, str]] = []
    seen_names = set()
    name_like = re.compile(r".*", re.I)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if (
            isinstance(v, str)
            and v.strip()
            and not v.strip().lower().startswith('use these formulas')
            and v.strip() not in ('Team Total:', 'Date', 'Event')
            and (name_like.search(v.strip()) is not None)
        ):
            base_name = v.strip()
            # De-duplicate names by appending a counter suffix for repeats (e.g., 'TBA (2)')
            if base_name not in seen_names:
                seen_names.add(base_name)
                name = base_name
            else:
                # count existing occurrences
                n = 2
                name = f"{base_name} ({n})"
                while name in seen_names:
                    n += 1
                    name = f"{base_name} ({n})"
                seen_names.add(name)
            name_cols.append((c, name))

    # Determine which column actually holds status strings for each person (c or c+1)
    status_cols: Dict[str, int] = {}
    status_values = set(STATUS_MAP.keys())
    max_scan_rows = min(ws.max_row, 200)  # quick scan for detection
    for c, name in name_cols:
        count_c = 0
        count_c1 = 0
        for r in range(3, max_scan_rows + 1):
            v_c = ws.cell(r, c).value
            if isinstance(v_c, str) and v_c in status_values:
                count_c += 1
            v_c1 = ws.cell(r, c + 1).value if c + 1 <= ws.max_column else None
            if isinstance(v_c1, str) and v_c1 in status_values:
                count_c1 += 1
        status_cols[name] = c if count_c >= count_c1 else (c + 1)

    # Gather all sessions (rows with a Date and Event)
    sessions: List[OrderedDict] = []
    session_rows: List[Tuple[int, str, datetime]] = []  # (row, eventId, date)
    event_labels: List[str] = []

    for r in range(3, ws.max_row + 1):
        date_val = ws.cell(r, 2).value
        event_label = ws.cell(r, 6).value
        if isinstance(date_val, datetime) and isinstance(event_label, str) and event_label.strip():
            eid = event_id(event_label)
            dstr = date_val.strftime('%Y-%m-%d')
            dow = dow_from_date(date_val)
            sid = f"{dstr}_{eid}"
            sessions.append(OrderedDict([
                ('id', sid),
                ('date', dstr),
                ('dow', dow),
                ('eventTypeId', eid),
            ]))
            session_rows.append((r, eid, date_val))
            event_labels.append(event_label)

    # People list
    people: List[OrderedDict] = []
    for _, name in name_cols:
        pid = slugify_id(name)
        people.append(OrderedDict([
            ('id', pid),
            ('displayName', name),
            ('active', True),
            ('tags', []),
            ('serviceDays', []),  # fill below
        ]))

    # Build serviceDays by scanning records and collecting days where not non_service/None
    service_days_by_person: Dict[str, set] = {p['id']: set() for p in people}

    # Records
    records: List[OrderedDict] = []
    rid = 1
    for (r, eid, d) in session_rows:
        for _, name in name_cols:
            col = status_cols[name]
            raw = ws.cell(r, col).value
            status: Optional[str]
            if isinstance(raw, str) and raw.strip():
                status = STATUS_MAP.get(raw.strip(), None)
            else:
                status = None  # missing

            pid = slugify_id(name)
            sid = f"{d.strftime('%Y-%m-%d')}_{eid}"
            records.append(OrderedDict([
                ('id', f"r_{rid}"),
                ('sessionId', sid),
                ('personId', pid),
                ('status', status),
            ]))
            rid += 1

            # Track service day if status is an actual attendance (not non_service/missing)
            if status and status not in ('non_service',):
                abbr = DAY_ABBR[d.weekday()]
                service_days_by_person[pid].add(abbr)

    # Apply serviceDays to people, preserving order
    for p in people:
        abbrs = ['Mon','Tue','Wed','Thu','Fri']  # Only include weekdays by default
        got = [d for d in abbrs if d in service_days_by_person[p['id']]]
        p['serviceDays'] = got

    # Event types (unique)
    unique_events = []
    seen = set()
    for lbl in event_labels:
        if lbl not in seen:
            seen.add(lbl)
            unique_events.append(lbl)
    eventTypes: List[OrderedDict] = []
    for lbl in unique_events:
        eid = event_id(lbl)
        weight = 0.95 if eid == 'meeting' else 1.0
        eventTypes.append(OrderedDict([
            ('id', eid),
            ('label', lbl),
            ('weight', weight),
        ]))

    # Settings (use template defaults)
    settings = [OrderedDict([
        ('id', 'app'),
        ('teamName', 'Attendance'),
        ('tardyThresholdMins', 5),
        ('legendThresholds', OrderedDict([
            ('low', 0.75),
            ('mid', 0.89),
            ('high', 0.9),
        ])),
    ])]

    data = OrderedDict([
        ('people', people),
        ('eventTypes', eventTypes),
        ('sessions', sessions),
        ('records', records),
        ('settings', settings),
    ])
    return data


def convert_all():
    root = Path.cwd()
    input_dir = root / 'JSON' / 'Attendance Roster Tracker'
    assert input_dir.exists(), f"Input dir not found: {input_dir}"

    xlsx_files = sorted(input_dir.glob('Attendance Roster - Team *.xlsx'))
    outputs: List[Tuple[str, str, str]] = []  # (source, output, status)

    for x in xlsx_files:
        if x.name.endswith('Team Helaman.xlsx'):
            continue  # skip template
        try:
            team = x.stem.replace('Attendance Roster - Team ', '').strip()
            out_name = f"Attendance{team}.json"
            out_path = root / 'JSON' / out_name
            data = parse_attendance_sheet(x)
            with out_path.open('w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            outputs.append((x.name, out_name, 'Success'))
        except Exception as e:
            outputs.append((x.name, '', f"Error: {e}"))

    # Print summary TSV to stdout for easy capture
    print('Source\tOutput\tStatus')
    for row in outputs:
        print('\t'.join(row))


if __name__ == '__main__':
    convert_all()
